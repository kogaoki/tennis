import streamlit as st
import pandas as pd
import itertools
import requests
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
from PyPDF2 import PdfReader, PdfWriter
import tempfile
import pathlib


st.set_page_config(layout="wide")
st.title("大会運営システム：リーグ対戦表＆スコアシート生成")

st.sidebar.header("設定")
total_pairs = st.sidebar.number_input("総ペア数", min_value=2, max_value=100, value=13, step=1)
pairs_per_league = st.sidebar.selectbox("1リーグに入れるペア数", options=[2, 3, 4, 5], index=2)
court_count = st.sidebar.number_input("使用コート数（進行表用）", min_value=1, max_value=10, value=2, step=1)

base_league_count = total_pairs // pairs_per_league
remainder = total_pairs % pairs_per_league

st.sidebar.markdown(f"**→ {base_league_count}リーグ + {remainder}ペア余り**")

extra_league_targets = []
if remainder > 0:
    st.sidebar.write("### 余りの振り分け先")
    options = [f"{chr(ord('A') + i)}" for i in range(base_league_count)]
    extra_league_targets = st.sidebar.multiselect(
        "追加するリーグを選択（上から順がおすすめ）",
        options,
        default=options[:remainder],
        max_selections=remainder
    )
    if len(extra_league_targets) != remainder:
        st.stop()

league_assignments = []
for i in range(base_league_count):
    league_size = pairs_per_league + (1 if chr(ord('A') + i) in extra_league_targets else 0)
    league_name = chr(ord('A') + i)
    league_assignments.append([f"{league_name}{j+1}" for j in range(league_size)])

st.write("### リーグごとの選手情報入力")
league_pair_data = {}

for i, league in enumerate(league_assignments):
    league_name = chr(ord('A') + i)
    st.subheader(f"{league_name}リーグ 選手入力")
    df = pd.DataFrame({
        "ペア番号": league,
        "所属": ["" for _ in league],
        "選手1": ["" for _ in league],
        "選手2": ["" for _ in league]
    })
    with st.container():
        try:
            edited = st.data_editor(
                df,
                column_config={"ペア番号": st.column_config.TextColumn(disabled=True)},
                use_container_width=True,
                hide_index=True,
                key=f"editor_{league_name}"
            )
            league_pair_data[league_name] = edited
        except Exception as e:
            st.error(f"{league_name}リーグの入力中にエラーが発生しました: {e}")
            continue

st.write("### リーグ対戦表のExcel出力")
if st.button("Excelダウンロード用にエクスポート"):
    wb = Workbook()
    ws = wb.active
    ws.title = "全リーグまとめ"
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold = Font(bold=True)
    current_row = 1

    for league_name, df in league_pair_data.items():
        if df.empty:
            continue
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=1, value=f"{league_name}リーグ").alignment = center
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1

        headers = ["No", "ペア名", "チーム名"] + [str(i + 1) for i in range(len(df))] + ["順位"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.alignment = center
            cell.font = bold
            cell.border = border
        current_row += 1

        for i, row in df.iterrows():
            pair = row["ペア番号"]
            team = row["所属"]
            name = f"{row['選手1']}・{row['選手2']}" if row["選手2"] else row["選手1"]
            row_data = [i + 1, name, team] + ["/" if j == i else "" for j in range(len(df))] + [""]
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.alignment = center
                cell.border = border
            current_row += 1
        current_row += 1

    output = BytesIO()
    wb.save(output)
    st.download_button("リーグ対戦表（Excel）をダウンロード", output.getvalue(), file_name="リーグ対戦表.xlsx")

# スコアシートPDF出力
if st.button("スコアシートPDFをダウンロード"):
    try:
        font_url = "https://raw.githubusercontent.com/kogaoki/tennis/main/NotoSansJP-VariableFont_wght.ttf"
        font_response = requests.get(font_url)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".ttf") as tmp_font_file:
            tmp_font_file.write(font_response.content)
            tmp_font_file.flush()
            font_path = tmp_font_file.name

        pdfmetrics.registerFont(TTFont("CustomJP", font_path))

        coords = {
            "no1": (92, 188), "team1": (213, 188), "p1_1": (187, 221), "p1_2": (187, 257),
            "no2": (361, 187), "team2": (477, 187), "p2_1": (453, 221), "p2_2": (452, 257)
        }

        def get_info(code):
            for league_df in league_pair_data.values():
                row = league_df[league_df["ペア番号"] == code]
                if not row.empty:
                    team = row.iloc[0]["所属"]
                    p1 = row.iloc[0]["選手1"]
                    p2 = row.iloc[0]["選手2"]
                    return team, p1, p2
            return "", "", ""

        match_schedule = []
        for league_name, df in league_pair_data.items():
            pairs = df["ペア番号"].tolist()
            if len(pairs) == 3:
                ordered = [(pairs[0], pairs[1]), (pairs[0], pairs[2]), (pairs[1], pairs[2])]
            elif len(pairs) == 4:
                ordered = [(pairs[0], pairs[1]), (pairs[2], pairs[3]), (pairs[0], pairs[2]),
                           (pairs[1], pairs[3]), (pairs[0], pairs[3]), (pairs[1], pairs[2])]
            else:
                ordered = list(itertools.combinations(pairs, 2))
            for m in ordered:
                match_schedule.append({"リーグ": league_name, "ペア1": m[0], "ペア2": m[1]})

        if not match_schedule:
            st.warning("対戦カードが作成されていません。選手情報を入力してください。")
            st.stop()

        # 背景テンプレートPDF
        template_url = "https://raw.githubusercontent.com/kogaoki/tennis/main/scoresheet.pdf"
        template_response = requests.get(template_url)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_template_file:
            tmp_template_file.write(template_response.content)
            tmp_template_path = tmp_template_file.name

        output = PdfWriter()

        for match in match_schedule:
            team1, p1_1, p1_2 = get_info(match["ペア1"])
            team2, p2_1, p2_2 = get_info(match["ペア2"])

            # 一時PDFに文字だけ描画
            overlay_buffer = BytesIO()
            c = canvas.Canvas(overlay_buffer, pagesize=A4)
            height = A4[1]

            def draw_text(x, y, text, center=False, fontsize=9):
                c.setFont("CustomJP", fontsize)
                if center:
                    text_width = c.stringWidth(text, "CustomJP", fontsize)
                    x -= text_width / 2
                c.drawString(x, height - y, text)

            draw_text(coords["team1"][0] + 5, coords["team1"][1], team1, center=True, fontsize=8)
            draw_text(coords["p1_1"][0], coords["p1_1"][1], p1_1, fontsize=12)
            if p1_2:
                draw_text(coords["p1_2"][0], coords["p1_2"][1], p1_2, fontsize=12)
            draw_text(coords["no1"][0], coords["no1"][1], match["ペア1"])

            draw_text(coords["team2"][0] + 5, coords["team2"][1], team2, center=True, fontsize=8)
            draw_text(coords["p2_1"][0], coords["p2_1"][1], p2_1, fontsize=12)
            if p2_2:
                draw_text(coords["p2_2"][0], coords["p2_2"][1], p2_2, fontsize=12)
            draw_text(coords["no2"][0], coords["no2"][1], match["ペア2"])

            c.showPage()
            c.save()
            overlay_buffer.seek(0)

            template_reader = PdfReader(tmp_template_path)
            overlay_reader = PdfReader(overlay_buffer)
            template_page = template_reader.pages[0]
            overlay_page = overlay_reader.pages[0]
            template_page.merge_page(overlay_page)
            output.add_page(template_page)

        final_buffer = BytesIO()
        output.write(final_buffer)
        final_buffer.seek(0)

        st.download_button("PDFスコアシートをダウンロード", final_buffer, file_name="score_sheets.pdf", mime="application/pdf")

    except Exception as e:
        st.error(f"PDF出力中にエラーが発生しました: {e}")
